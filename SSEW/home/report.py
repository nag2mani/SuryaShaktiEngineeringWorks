import pytz
from .forms import *
from .models import *
from io import BytesIO
from openpyxl import Workbook
from datetime import datetime
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.shortcuts import render


def fund_expense_list(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    id = request.GET.get('id')

    if from_date and to_date:
        from_date = parse_date(from_date)
        to_date = parse_date(to_date)

        if id == 'fund_expense_list':
            expenses = Expenses.objects.filter(date_time__range=[from_date, to_date])
            funds = FundIn.objects.filter(date_time__range=[from_date, to_date])
            fund_expense_list = list(expenses) + list(funds)  # Combine querysets into a list
            return render(request, 'fund_expense_list.html', {'fund_expense_list': fund_expense_list})
    
    return render(request, 'fund_expense_list.html', {'fund_expense_list': []})


def employee_list(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    id = request.GET.get('id')

    if from_date and to_date:
        from_date = parse_date(from_date)
        to_date = parse_date(to_date)

        if id == 'employee_list':
            employees = Employee.objects.filter(date_time__range=[from_date, to_date])
            return render(request, 'employee_list.html', {'employees': employees})

    return render(request, 'employee_list.html', {'employees': []})


def inventory_list(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    id = request.GET.get('id')

    if from_date and to_date:
        from_date = parse_date(from_date)
        to_date = parse_date(to_date)

        if id == 'inventory_list':
            inventories = Inventory.objects.filter(date_time__range=[from_date, to_date])
            return render(request, 'inventory_list.html', {'inventories': inventories})

    return render(request, 'inventory_list.html', {'inventories': []})


def generate_fund_expense_report_xlsx(from_date, to_date):
    # Filter data based on the date range
    expenses = Expenses.objects.filter(date_time__range=[from_date, to_date])
    funds = FundIn.objects.filter(date_time__range=[from_date, to_date])

    # Create a workbook and add a worksheet.
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Fund Expense Report"

    # Define the headers based on model fields
    expense_headers = [field.name for field in Expenses._meta.fields]
    fund_headers = [field.name for field in FundIn._meta.fields]

    # Write headers for expenses
    for col_num, header in enumerate(expense_headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Write expense data
    row = 2
    total_cash_expenses = 0
    total_online_expenses = 0
    for expense in expenses:
        for col_num, field in enumerate(expense_headers, 1):
            value = getattr(expense, field)
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            if field in ['category', 'subcategory', 'employee'] and value:
                value = str(value)
            worksheet.cell(row=row, column=col_num, value=value)
            if field == 'cash_expenses':
                total_cash_expenses += value
            if field == 'online_expenses':
                total_online_expenses += value
        row += 1

    # Leave a row blank
    row += 3

    # Write headers for fund in
    for col_num, header in enumerate(fund_headers, 1):
        worksheet.cell(row=row, column=col_num, value=header)

    # Write fund in data
    row += 1
    total_cash_fund = 0
    total_online_fund = 0
    for fund in funds:
        for col_num, field in enumerate(fund_headers, 1):
            value = getattr(fund, field)
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            if field in ['source', 'subsource', 'received_by'] and value:
                value = str(value)
            worksheet.cell(row=row, column=col_num, value=value)
            if field == 'cash_fund':
                total_cash_fund += value
            if field == 'online_fund':
                total_online_fund += value
        row += 1

    # Calculate totals and profit
    total_expenses = total_cash_expenses + total_online_expenses
    total_fundin = total_cash_fund + total_online_fund
    profit = total_fundin - total_expenses

    # Write totals and profit in the worksheet
    row += 2
    worksheet.cell(row=row, column=1, value="Total Cash Expenses")
    worksheet.cell(row=row, column=2, value=total_cash_expenses)

    row += 1
    worksheet.cell(row=row, column=1, value="Total Online Expenses")
    worksheet.cell(row=row, column=2, value=total_online_expenses)

    row += 1
    worksheet.cell(row=row, column=1, value="Total Cash FundIn")
    worksheet.cell(row=row, column=2, value=total_cash_fund)

    row += 1
    worksheet.cell(row=row, column=1, value="Total Online FundIn")
    worksheet.cell(row=row, column=2, value=total_online_fund)

    row += 1
    worksheet.cell(row=row, column=1, value="Profit")
    worksheet.cell(row=row, column=2, value=profit)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Create the HttpResponse
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fund_expense_report.xlsx'

    return response




def generate_employee_report_xlsx(from_date, to_date):
    # Filter data based on the date range
    employees = Employee.objects.filter(date_time__range=[from_date, to_date])

    # Create a workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Employee Report"

    # Define the headers based on model fields
    headers = [field.name for field in Employee._meta.fields]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Write employee data
    row = 2
    for employee in employees:
        for col_num, field in enumerate(headers, 1):
            value = getattr(employee, field)
            if isinstance(value, (pytz.datetime.datetime, pytz.datetime.time)):
                value = value.replace(tzinfo=None)
            worksheet.cell(row=row, column=col_num, value=value)
        row += 1

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Create the HttpResponse
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_report.xlsx'

    return response




def generate_inventory_report_xlsx(from_date, to_date):
    # Filter data based on the date range
    inventories = Inventory.objects.filter(date_time__range=[from_date, to_date])

    # Create a workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventory Report"

    # Define the headers based on model fields
    headers = [field.name for field in Inventory._meta.fields]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Write inventory data
    row = 2
    for inventory in inventories:
        for col_num, field in enumerate(headers, 1):
            value = getattr(inventory, field)
            if isinstance(value, (pytz.datetime.datetime, pytz.datetime.time)):
                value = value.replace(tzinfo=None)
            worksheet.cell(row=row, column=col_num, value=value)
        row += 1

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Create the HttpResponse
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventory_report.xlsx'

    return response



def generate_category_wise_report_xlsx(from_date, to_date, category_id):
    expenses = Expenses.objects.filter(date_time__range=[from_date, to_date])
    if category_id:
        expenses = expenses.filter(category_id=category_id)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=category_wise_report.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Expenses'

    columns = ['Date', 'Online Expenses', 'Cash Expenses', 'Name of Person', 'Remark']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=row_num, column=col_num, value=column_title)

    for expense in expenses:
        row_num += 1
        row = [
            expense.date_time,
            expense.online_expenses,
            expense.cash_expenses,
            expense.name_of_person,
            expense.remark
        ]
        for col_num, cell_value in enumerate(row, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(response)
    return response

def generate_source_wise_report_xlsx(from_date, to_date, source_id):
    fundins = FundIn.objects.filter(date_time__range=[from_date, to_date])
    if source_id:
        fundins = fundins.filter(source_id=source_id)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=source_wise_report.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'FundIns'

    columns = ['Date', 'Online Fund', 'Cash Fund', 'Name', 'Remark']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=row_num, column=col_num, value=column_title)

    for fundin in fundins:
        row_num += 1
        row = [
            fundin.date_time,
            fundin.online_fund,
            fundin.cash_fund,
            fundin.name,
            fundin.remark
        ]
        for col_num, cell_value in enumerate(row, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(response)
    return response
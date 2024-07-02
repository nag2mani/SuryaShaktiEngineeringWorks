from django.shortcuts import render, redirect
from django.utils.dateparse import parse_date
from django.http import JsonResponse, HttpResponse
from .models import *
from .forms import *

from openpyxl import Workbook
from io import BytesIO
import pytz
from datetime import datetime
#Instruction : Use and pass only singular variable name throught the application.

def home(request):
    total_cash_expense = Expenses.objects.aggregate(models.Sum('cash_expenses'))['cash_expenses__sum'] or 0
    total_online_expense = Expenses.objects.aggregate(models.Sum('online_expenses'))['online_expenses__sum'] or 0
    total_cash_fund = FundIn.objects.aggregate(models.Sum('cash_fund'))['cash_fund__sum'] or 0
    total_online_fund = FundIn.objects.aggregate(models.Sum('online_fund'))['online_fund__sum'] or 0
    total_fund = total_cash_fund + total_online_fund

    total_expense = total_cash_expense + total_online_expense

    total_cash_in_hand = total_cash_fund - total_cash_expense

    # Fetch total number of employees
    total_employees = Employee.objects.count()

    # Fetch recent cash transactions (expenses and fund-ins)
    recent_cash_expense = Expenses.objects.filter(cash_expenses__gt=0).order_by('-date_time')[:2]
    recent_cash_fund_ins = FundIn.objects.filter(cash_fund__gt=0).order_by('-date_time')[:3]

    # Fetch recent online transactions (expenses and fund-ins)
    recent_online_expense = Expenses.objects.filter(online_expenses__gt=0).order_by('-date_time')[:2]
    recent_online_fund_ins = FundIn.objects.filter(online_fund__gt=0).order_by('-date_time')[:3]

    context = {
        'total_expense': total_expense,
        'total_cash_in_hand': total_cash_in_hand,
        'total_employees': total_employees,
        'recent_cash_expense': recent_cash_expense,
        'recent_cash_fund_ins': recent_cash_fund_ins,
        'recent_online_expense': recent_online_expense,
        'recent_online_fund_ins': recent_online_fund_ins,
    }
    return render(request, 'index.html', context)

def load_subcategories(request):
    category_id = request.GET.get('category')
    subcategories = SubCategory.objects.filter(category_id=category_id).all()
    return JsonResponse(list(subcategories.values('id', 'name')), safe=False)

def load_subsources(request):
    source_id = request.GET.get('source')
    subsources = SubSource.objects.filter(source_id=source_id).all()
    return JsonResponse(list(subsources.values('id', 'name')), safe=False)

def add_expense(request):
    if request.method == 'POST':
        form = ExpensesForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')  # Redirect to the home
    else:
        form = ExpensesForm()
    return render(request, 'add_expense.html', {'form': form})


def add_fund(request):
    if request.method == 'POST':
        form = FundInForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')  # Redirect to the home
    else:
        form = FundInForm()
    return render(request, 'add_fundin.html', {'form': form})


def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')  # Redirect to the home
    else:
        form = EmployeeForm()
    return render(request, 'add_employee.html', {'form': form})


def add_inventory(request):
    if request.method == 'POST':
        form = InventoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')  # Redirect to the home
    else:
        form = InventoryForm()
    return render(request, 'add_inventory.html', {'form': form})


def all_data(request):
    expenses = Expenses.objects.all()
    fundin = FundIn.objects.all()
    employees = Employee.objects.all()
    inventory = Inventory.objects.all()

    context = {
        'expenses': expenses,
        'fundin': fundin,
        'employees': employees,
        'inventory': inventory,
    }
    return render(request, 'all_data.html', context)


def generate_report(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    report_type = request.GET.get('report_type')

    if from_date and to_date:
        from_date = parse_date(from_date)
        to_date = parse_date(to_date)

        if report_type == 'fund_expense_xlsx':
            return generate_fund_expense_report_xlsx(from_date, to_date)
        elif report_type == 'employee_xlsx':
            return generate_employee_report_xlsx(from_date, to_date)
        elif report_type == 'inventory_xlsx':
            return generate_inventory_report_xlsx(from_date, to_date)
        elif report_type == 'category_wise_report_xlsx':
            return generate_category_wise_report_xlsx(from_date, to_date)
    else:
        return render(request, 'report.html')


def fund_expense_list(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    id = request.GET.get('id')

    if from_date and to_date:
        from_date = parse_date(from_date)
        to_date = parse_date(to_date)

        if id == 'fund_expense_list':
            expenses = Expenses.objects.filter(date_time__range=[from_date, to_date])
            funds = FundIn.objects.filter(date_time__range=[from_date, to_date])
            fund_expense_list = list(expenses) + list(funds)  # Combine querysets into a list
            return render(request, 'fund_expense_list.html', {'fund_expense_list': fund_expense_list})
    
    return render(request, 'fund_expense_list.html', {'fund_expense_list': []})


def employee_list(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    id = request.GET.get('id')

    if from_date and to_date:
        from_date = parse_date(from_date)
        to_date = parse_date(to_date)

        if id == 'employee_list':
            employees = Employee.objects.filter(date_time__range=[from_date, to_date])
            return render(request, 'employee_list.html', {'employees': employees})

    return render(request, 'employee_list.html', {'employees': []})


def inventory_list(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    id = request.GET.get('id')

    if from_date and to_date:
        from_date = parse_date(from_date)
        to_date = parse_date(to_date)

        if id == 'inventory_list':
            inventories = Inventory.objects.filter(date_time__range=[from_date, to_date])
            return render(request, 'inventory_list.html', {'inventories': inventories})

    return render(request, 'inventory_list.html', {'inventories': []})


def generate_fund_expense_report_xlsx(from_date, to_date):
    # Filter data based on the date range
    expenses = Expenses.objects.filter(date_time__range=[from_date, to_date])
    funds = FundIn.objects.filter(date_time__range=[from_date, to_date])

    # Create a workbook and add a worksheet.
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Fund Expense Report"

    # Define the headers based on model fields
    expense_headers = [field.name for field in Expenses._meta.fields]
    fund_headers = [field.name for field in FundIn._meta.fields]

    # Write headers for expenses
    for col_num, header in enumerate(expense_headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Write expense data
    row = 2
    total_cash_expenses = 0
    total_online_expenses = 0
    for expense in expenses:
        for col_num, field in enumerate(expense_headers, 1):
            value = getattr(expense, field)
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            if field in ['category', 'subcategory', 'employee'] and value:
                value = str(value)
            worksheet.cell(row=row, column=col_num, value=value)
            if field == 'cash_expenses':
                total_cash_expenses += value
            if field == 'online_expenses':
                total_online_expenses += value
        row += 1

    # Leave a row blank
    row += 3

    # Write headers for fund in
    for col_num, header in enumerate(fund_headers, 1):
        worksheet.cell(row=row, column=col_num, value=header)

    # Write fund in data
    row += 1
    total_cash_fund = 0
    total_online_fund = 0
    for fund in funds:
        for col_num, field in enumerate(fund_headers, 1):
            value = getattr(fund, field)
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            if field in ['source', 'subsource', 'received_by'] and value:
                value = str(value)
            worksheet.cell(row=row, column=col_num, value=value)
            if field == 'cash_fund':
                total_cash_fund += value
            if field == 'online_fund':
                total_online_fund += value
        row += 1

    # Calculate totals and profit
    total_expenses = total_cash_expenses + total_online_expenses
    total_fundin = total_cash_fund + total_online_fund
    profit = total_fundin - total_expenses

    # Write totals and profit in the worksheet
    row += 2
    worksheet.cell(row=row, column=1, value="Total Cash Expenses")
    worksheet.cell(row=row, column=2, value=total_cash_expenses)

    row += 1
    worksheet.cell(row=row, column=1, value="Total Online Expenses")
    worksheet.cell(row=row, column=2, value=total_online_expenses)

    row += 1
    worksheet.cell(row=row, column=1, value="Total Cash FundIn")
    worksheet.cell(row=row, column=2, value=total_cash_fund)

    row += 1
    worksheet.cell(row=row, column=1, value="Total Online FundIn")
    worksheet.cell(row=row, column=2, value=total_online_fund)

    row += 1
    worksheet.cell(row=row, column=1, value="Profit")
    worksheet.cell(row=row, column=2, value=profit)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Create the HttpResponse
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=fund_expense_report.xlsx'

    return response




def generate_employee_report_xlsx(from_date, to_date):
    # Filter data based on the date range
    employees = Employee.objects.filter(date_time__range=[from_date, to_date])

    # Create a workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Employee Report"

    # Define the headers based on model fields
    headers = [field.name for field in Employee._meta.fields]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Write employee data
    row = 2
    for employee in employees:
        for col_num, field in enumerate(headers, 1):
            value = getattr(employee, field)
            if isinstance(value, (pytz.datetime.datetime, pytz.datetime.time)):
                value = value.replace(tzinfo=None)
            worksheet.cell(row=row, column=col_num, value=value)
        row += 1

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Create the HttpResponse
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_report.xlsx'

    return response




def generate_inventory_report_xlsx(from_date, to_date):
    # Filter data based on the date range
    inventories = Inventory.objects.filter(date_time__range=[from_date, to_date])

    # Create a workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventory Report"

    # Define the headers based on model fields
    headers = [field.name for field in Inventory._meta.fields]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Write inventory data
    row = 2
    for inventory in inventories:
        for col_num, field in enumerate(headers, 1):
            value = getattr(inventory, field)
            if isinstance(value, (pytz.datetime.datetime, pytz.datetime.time)):
                value = value.replace(tzinfo=None)
            worksheet.cell(row=row, column=col_num, value=value)
        row += 1

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Create the HttpResponse
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventory_report.xlsx'

    return response


def generate_category_wise_expense_report_xlsx(from_date, to_date, category_id):
    expenses = Expenses.objects.filter(date_time__range=[from_date, to_date])
    if category_id:
        expenses = expenses.filter(category_id=category_id)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Category-wise Expense Report"

    headers = [field.name for field in Expenses._meta.fields]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    row = 2
    for expense in expenses:
        for col_num, field in enumerate(headers, 1):
            value = getattr(expense, field)
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            if field in ['category', 'subcategory', 'employee'] and value:
                value = str(value)
            worksheet.cell(row=row, column=col_num, value=value)
        row += 1

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=category_wise_expense_report.xlsx'
    return response

def generate_source_wise_fundin_report_xlsx(from_date, to_date, source_id):
    fundins = FundIn.objects.filter(date_time__range=[from_date, to_date])
    if source_id:
        fundins = fundins.filter(source_id=source_id)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Source-wise Fund In Report"

    headers = [field.name for field in FundIn._meta.fields]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    row = 2
    for fundin in fundins:
        for col_num, field in enumerate(headers, 1):
            value = getattr(fundin, field)
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            if field in ['source', 'subsource', 'received_by'] and value:
                value = str(value)
            worksheet.cell(row=row, column=col_num, value=value)
        row += 1

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=source_wise_fundin_report.xlsx'
    return response